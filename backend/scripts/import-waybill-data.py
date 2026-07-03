#!/usr/bin/env python3
import csv
import json
import os
import subprocess
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_DIR = ROOT / "split_result"
DEFAULT_EMPTY_REPORT = ROOT / "import_reports" / "waybill_empty_rows.csv"
TABLE = "dwd.waybill_detail"
INCLUDE_RAW_ROW = "--include-raw-row" in sys.argv
RESUME = "--resume" in sys.argv
TRUNCATE = "--truncate" in sys.argv

FIELDS = [
    ("waybill_no", "运单编号", "text"),
    ("customer_name", "客户名称", "text"),
    ("shipping_time", "寄件时间", "timestamp"),
    ("shipping_finance_center", "寄件财务中心", "text"),
    ("shipping_station", "寄件网点", "text"),
    ("origin_city", "始发地", "text"),
    ("destination_station", "目的网点", "text"),
    ("destination_city", "目的地", "text"),
    ("piece_count", "件数", "numeric"),
    ("settlement_method", "结算方式", "text"),
    ("total_freight", "总运费", "numeric"),
    ("waybill_freight", "运单运费", "numeric"),
    ("insurance_fee", "保价费", "numeric"),
    ("insured_amount", "保价金额", "numeric"),
    ("freight_collect_amount", "到付运费", "numeric"),
    ("standard_price_amount", "标准价金额", "numeric"),
    ("other_receivable_fee", "应收其他费", "numeric"),
    ("manual_fee", "手工费", "numeric"),
    ("loaded_weight", "装载重量", "numeric"),
    ("entry_time", "录入时间", "timestamp"),
    ("waybill_status", "运单状态", "text"),
    ("goods_name", "物品名称", "text"),
    ("package_volume_weight", "包裹体积重", "numeric"),
    ("package_total_volume", "包裹总体积", "numeric"),
    ("package_billing_weight", "包裹计费重量", "numeric"),
    ("actual_weight", "实际重量", "numeric"),
    ("internal_billing_weight", "内部计费重量", "numeric"),
    ("order_weight", "订单重量", "numeric"),
    ("pickup_station_code", "揽件网点编号", "text"),
    ("pickup_station_name", "揽件网点", "text"),
    ("pickup_time", "揽件时间", "timestamp"),
    ("dispatch_courier_code", "派件业务员编号", "text"),
    ("dispatch_courier_name", "派件业务员", "text"),
    ("dispatch_time", "派件时间", "timestamp"),
    ("dispatch_station_code", "派件网点编号", "text"),
    ("sign_flag", "签收标识", "boolean"),
    ("sign_station_code", "签收网点编号", "text"),
    ("sign_station_name", "签收网点", "text"),
    ("sign_time", "签收时间", "timestamp"),
    ("customer_order_no", "客户订单编号", "text"),
    ("waybill_source_code", "运单来源编号", "text"),
    ("waybill_source", "运单来源", "text"),
    ("order_source_code", "订单来源编号", "text"),
    ("order_source", "订单来源", "text"),
    ("shipping_method_code", "寄件方式编号", "text"),
    ("shipping_method", "寄件方式", "text"),
    ("dispatch_method_code", "派件方式编码", "text"),
    ("dispatch_method", "派件方式", "text"),
    ("void_flag_code", "作废标记编码", "text"),
    ("void_flag", "作废标记", "text"),
    ("return_flag_code", "转退件编码", "text"),
    ("return_flag", "转退件标识", "text"),
    ("has_return_receipt", "是否签回单", "boolean"),
    ("return_receipt_no", "回单/回单原单号", "text"),
    ("customer_code", "客户编号", "text"),
    ("sender_name", "寄件人姓名", "text"),
    ("sender_phone", "寄件人手机号", "text"),
    ("sender_country", "寄件国家", "text"),
    ("sender_province", "寄件省份", "text"),
    ("sender_city", "寄件城市", "text"),
    ("sender_area", "寄件区域", "text"),
    ("sender_address", "寄件详细地址", "text"),
    ("shipping_station_code", "寄件网点编号", "text"),
    ("shipping_finance_center_code", "寄件财务中心编号", "text"),
    ("dispatch_finance_center_code", "派件财务中心编号", "text"),
    ("dispatch_finance_center", "派件财务中心", "text"),
    ("product_type_code", "产品类型编号", "text"),
    ("product_type", "产品类型", "text"),
    ("insurance_required_code", "是否需要保价编码", "text"),
    ("insurance_required", "是否需要保价", "boolean"),
    ("cod_flag", "代收货款标记", "boolean"),
    ("cod_amount", "代收货款金额", "numeric"),
    ("cod_fee", "代收货款手续费", "numeric"),
    ("receiver_name", "收件人姓名", "text"),
    ("receiver_company", "收件公司", "text"),
    ("receiver_phone", "收件人手机号", "text"),
    ("receiver_tel", "收件人座机", "text"),
    ("profit", "利润", "numeric"),
    ("market_price", "市场价", "numeric"),
    ("receiver_country", "收件国家", "text"),
    ("receiver_province", "收件省份", "text"),
    ("receiver_city", "收件城市", "text"),
    ("receiver_area", "收件区域", "text"),
    ("receiver_address", "收件详细地址", "text"),
    ("customer_owner_station", "客户归属网点", "text"),
    ("cross_code", "十字码", "text"),
    ("estimated_time", "预估时效", "timestamp"),
    ("unit_kg_price", "单公斤价格", "numeric"),
    ("global_discount", "全局折扣", "numeric"),
    ("station_discount", "网点折扣", "numeric"),
    ("standard_quote", "标准报价", "numeric"),
    ("receipt_type_code", "回单类型", "text"),
    ("receipt_type_name", "回单类型名称", "text"),
    ("warehouse_entry_fee", "入仓费", "numeric"),
    ("receipt_fee", "回单费", "numeric"),
    ("sender_town", "寄件乡镇", "text"),
    ("sender_street", "寄件街道", "text"),
    ("sender_town_id", "寄件乡镇ID", "text"),
    ("receiver_town", "收件乡镇", "text"),
    ("receiver_street", "收件街道", "text"),
    ("receiver_town_id", "收件乡镇ID", "text"),
]


def load_dotenv() -> None:
    env_path = ROOT / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key, value)


def env(name: str, default: str) -> str:
    return os.environ.get(name, default)


def psql_env() -> dict[str, str]:
    values = os.environ.copy()
    values["PGPASSWORD"] = env("DATABASE_PASSWORD", "chatbi_password")
    return values


def psql_base() -> list[str]:
    return [
        "psql",
        "-h",
        env("DATABASE_HOST", "localhost"),
        "-p",
        env("DATABASE_PORT", "5432"),
        "-U",
        env("DATABASE_USER", "chatbi_user"),
        "-d",
        env("DATABASE_NAME", "chatbi_db"),
        "-v",
        "ON_ERROR_STOP=1",
    ]


def run_sql(sql: str) -> None:
    subprocess.run(psql_base(), input=sql, text=True, check=True, env=psql_env())


def scalar_sql(sql: str) -> str:
    result = subprocess.run(
        psql_base() + ["-Atc", sql],
        text=True,
        check=True,
        capture_output=True,
        env=psql_env(),
    )
    return result.stdout.strip()


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def parse_numeric(value) -> str:
    text = clean_text(value)
    if text in {"", "-", "—"}:
        return ""
    try:
        return str(Decimal(text))
    except (InvalidOperation, ValueError):
        return ""


def parse_time(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat(sep=" ")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).isoformat(sep=" ")
    text = clean_text(value)
    if text in {"", "-", "—"}:
        return ""
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt).isoformat(sep=" ")
        except ValueError:
            pass
    return ""


def parse_bool(value) -> str:
    text = clean_text(value).lower()
    if text in {"", "-", "—"}:
        return ""
    if text in {"是", "yes", "y", "true", "1", "已签收"}:
        return "true"
    if text in {"否", "no", "n", "false", "0", "未签收"}:
        return "false"
    return ""


def normalize(value, kind: str) -> str:
    if kind == "numeric":
        return parse_numeric(value)
    if kind == "timestamp":
        return parse_time(value)
    if kind == "boolean":
        return parse_bool(value)
    return clean_text(value)


def is_empty_row(row) -> bool:
    return not any(cell is not None and str(cell).strip() != "" for cell in row)


def raw_row_json(headers, values) -> str:
    payload = {}
    for header, value in zip(headers, values):
        if isinstance(value, datetime):
            payload[header] = value.replace(tzinfo=None).isoformat(sep=" ")
        elif isinstance(value, date):
            payload[header] = value.isoformat()
        elif isinstance(value, Decimal):
            payload[header] = str(value)
        else:
            payload[header] = value
    return json.dumps(payload, ensure_ascii=False)


def table_ddl() -> str:
    field_sql = []
    for name, _, kind in FIELDS:
        if kind == "numeric":
            typ = "NUMERIC"
        elif kind == "timestamp":
            typ = "TIMESTAMP"
        elif kind == "boolean":
            typ = "BOOLEAN"
        else:
            typ = "TEXT"
        field_sql.append(f"  {name} {typ}")
    fields = ",\n".join(field_sql)
    return f"""
CREATE SCHEMA IF NOT EXISTS dwd;

CREATE TABLE IF NOT EXISTS {TABLE} (
  source_file TEXT NOT NULL,
  source_sheet TEXT NOT NULL,
  source_row_no INT NOT NULL,
  query_scope TEXT,
{fields},
  raw_row JSONB,
  imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_waybill_detail_waybill_no ON {TABLE}(waybill_no);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_customer_order_no ON {TABLE}(customer_order_no);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_entry_time ON {TABLE}(entry_time);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_shipping_time ON {TABLE}(shipping_time);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_pickup_time ON {TABLE}(pickup_time);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_dispatch_time ON {TABLE}(dispatch_time);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_sign_time ON {TABLE}(sign_time);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_order_source ON {TABLE}(order_source);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_customer_code ON {TABLE}(customer_code);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_shipping_station ON {TABLE}(shipping_station_code);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_dispatch_station ON {TABLE}(dispatch_station_code);
CREATE INDEX IF NOT EXISTS idx_waybill_detail_status ON {TABLE}(waybill_status);
ALTER TABLE {TABLE} ALTER COLUMN raw_row DROP NOT NULL;
"""


def expected_headers() -> list[str]:
    return [header for _, header, _ in FIELDS]


def normalized_row(source_file: str, sheet_name: str, row_no: int, query_scope: str, headers, row):
    values = list(row) + [None] * (len(FIELDS) - len(row))
    values = values[: len(FIELDS)]
    output = [source_file, sheet_name, str(row_no), query_scope]
    for value, (_, _, kind) in zip(values, FIELDS):
        output.append(normalize(value, kind))
    if INCLUDE_RAW_ROW:
        output.append(raw_row_json(headers, values))
    return output


def import_file(path: Path, empty_writer) -> tuple[int, int]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    total_imported = 0
    total_empty = 0
    headers_expected = expected_headers()
    copy_cols = ["source_file", "source_sheet", "source_row_no", "query_scope"] + [
        name for name, _, _ in FIELDS
    ]
    if INCLUDE_RAW_ROW:
        copy_cols.append("raw_row")
    copy_sql = (
        f"\\copy {TABLE} ({', '.join(copy_cols)}) "
        "FROM STDIN WITH (FORMAT csv, NULL '', QUOTE '\"')"
    )

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            scope_row = next(rows)
            header_row = next(rows)
        except StopIteration:
            continue
        query_scope = clean_text(scope_row[0])
        headers = [clean_text(v) for v in list(header_row)[: len(FIELDS)]]
        if headers != headers_expected:
            raise RuntimeError(f"Unexpected headers in {path.name}/{sheet_name}: {headers}")

        print(f"Importing {path.name} / {sheet_name}...", flush=True)
        proc = subprocess.Popen(
            psql_base() + ["-c", copy_sql],
            stdin=subprocess.PIPE,
            text=True,
            env=psql_env(),
        )
        assert proc.stdin is not None
        writer = csv.writer(proc.stdin, lineterminator="\n")
        imported = 0
        empty = 0
        for row_no, row in enumerate(rows, start=3):
            if is_empty_row(row):
                empty_writer.writerow([path.name, sheet_name, row_no])
                empty += 1
                continue
            writer.writerow(normalized_row(path.name, sheet_name, row_no, query_scope, headers, row))
            imported += 1
            if imported % 100000 == 0:
                print(f"  {path.name}: imported {imported}, empty {empty}...", flush=True)
        proc.stdin.close()
        code = proc.wait()
        if code != 0:
            raise RuntimeError(f"COPY failed for {path.name}/{sheet_name}")
        print(f"  {path.name}: imported {imported}, empty {empty}", flush=True)
        total_imported += imported
        total_empty += empty
    workbook.close()
    return total_imported, total_empty


def expected_data_rows(path: Path) -> int:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    total = 0
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        total += max((sheet.max_row or 0) - 2, 0)
    workbook.close()
    return total


def imported_rows_for_file(path: Path) -> int:
    escaped = path.name.replace("'", "''")
    value = scalar_sql(f"SELECT COUNT(*) FROM {TABLE} WHERE source_file = '{escaped}';")
    return int(value or "0")


def delete_file_rows(path: Path) -> None:
    escaped = path.name.replace("'", "''")
    run_sql(f"DELETE FROM {TABLE} WHERE source_file = '{escaped}';")


def main() -> None:
    load_dotenv()
    args = [
        arg
        for arg in sys.argv[1:]
        if arg not in {"--truncate", "--resume", "--include-raw-row"}
    ]
    input_dir = Path(args[0]) if args else DEFAULT_INPUT_DIR
    empty_report = Path(args[1]) if len(args) > 1 else DEFAULT_EMPTY_REPORT
    files = sorted(input_dir.glob("寄件运单查询新*.xlsx"))
    if not files:
        raise SystemExit(f"No files matched in {input_dir}")

    print(f"Creating table {TABLE} if needed...")
    run_sql(table_ddl())
    if TRUNCATE:
        print(f"Truncating {TABLE}...")
        run_sql(f"TRUNCATE TABLE {TABLE};")

    empty_report.parent.mkdir(parents=True, exist_ok=True)
    before = scalar_sql(f"SELECT COUNT(*) FROM {TABLE};")
    print(f"Rows before import: {before}")
    if int(before or "0") > 0 and not TRUNCATE and not RESUME:
        raise SystemExit(
            f"{TABLE} already has {before} rows. Use --resume to continue safely "
            "or --truncate to reload from scratch."
        )
    total_imported = 0
    total_empty = 0
    with empty_report.open("w", newline="", encoding="utf-8") as f:
        empty_writer = csv.writer(f)
        empty_writer.writerow(["source_file", "source_sheet", "source_row_no"])
        for path in files:
            if RESUME and not TRUNCATE:
                expected = expected_data_rows(path)
                imported_existing = imported_rows_for_file(path)
                if imported_existing == expected:
                    print(f"Skipping {path.name}: already imported {imported_existing}/{expected}")
                    continue
                if imported_existing:
                    print(
                        f"Reimporting {path.name}: existing {imported_existing}, expected {expected}. "
                        "Deleting existing rows for this file first."
                    )
                    delete_file_rows(path)
            imported, empty = import_file(path, empty_writer)
            total_imported += imported
            total_empty += empty
    after = scalar_sql(f"SELECT COUNT(*) FROM {TABLE};")
    print(f"Imported rows: {total_imported}")
    print(f"Empty rows excluded: {total_empty}")
    print(f"Empty-row report: {empty_report}")
    print(f"Rows after import: {after}")


if __name__ == "__main__":
    main()
