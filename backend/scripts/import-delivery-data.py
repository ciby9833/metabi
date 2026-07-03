#!/usr/bin/env python3
import csv
import json
import os
import subprocess
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "split_result" / "派送数据.xlsx"
DEFAULT_EMPTY_REPORT = ROOT / "import_reports" / "delivery_empty_rows.csv"
TABLE = "dwd.delivery_event_detail"

EXPECTED_HEADERS = [
    "运单号",
    "经度",
    "纬度",
    "收件人姓名",
    "收件人手机号",
    "收件人省市区",
    "收件详细地址",
    "寄件人姓名",
    "寄件人电话",
    "寄件人省市区",
    "寄件人详细地址",
    "派件网点编码",
    "派件网点名称",
    "派件业务员编码",
    "派件业务员名称",
    "派件时间",
    "签收时间",
    "是否cod",
    "订单来源",
    "目的网点到件扫描时间",
]

COLUMNS = [
    "source_file",
    "source_sheet",
    "source_row_no",
    "waybill_no",
    "longitude",
    "latitude",
    "receiver_name",
    "receiver_phone",
    "receiver_region",
    "receiver_address",
    "sender_name",
    "sender_phone",
    "sender_region",
    "sender_address",
    "dispatch_station_code",
    "dispatch_station_name",
    "courier_code",
    "courier_name",
    "dispatch_time",
    "sign_time",
    "is_cod",
    "order_source",
    "destination_arrival_scan_time",
    "has_destination_arrival",
    "has_dispatch",
    "has_sign",
    "arrival_to_dispatch_minutes",
    "dispatch_to_sign_minutes",
    "arrival_to_sign_minutes",
    "process_status",
    "raw_row",
]

DDL = f"""
CREATE SCHEMA IF NOT EXISTS dwd;

CREATE TABLE IF NOT EXISTS {TABLE} (
  source_file TEXT NOT NULL,
  source_sheet TEXT NOT NULL,
  source_row_no INT NOT NULL,
  waybill_no TEXT,
  longitude NUMERIC(12, 8),
  latitude NUMERIC(12, 8),
  receiver_name TEXT,
  receiver_phone TEXT,
  receiver_region TEXT,
  receiver_address TEXT,
  sender_name TEXT,
  sender_phone TEXT,
  sender_region TEXT,
  sender_address TEXT,
  dispatch_station_code TEXT,
  dispatch_station_name TEXT,
  courier_code TEXT,
  courier_name TEXT,
  dispatch_time TIMESTAMP,
  sign_time TIMESTAMP,
  is_cod BOOLEAN,
  order_source TEXT,
  destination_arrival_scan_time TIMESTAMP,
  has_destination_arrival BOOLEAN NOT NULL DEFAULT false,
  has_dispatch BOOLEAN NOT NULL DEFAULT false,
  has_sign BOOLEAN NOT NULL DEFAULT false,
  arrival_to_dispatch_minutes NUMERIC(14, 2),
  dispatch_to_sign_minutes NUMERIC(14, 2),
  arrival_to_sign_minutes NUMERIC(14, 2),
  process_status TEXT,
  raw_row JSONB NOT NULL,
  imported_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_delivery_event_detail_waybill
  ON {TABLE}(waybill_no);
CREATE INDEX IF NOT EXISTS idx_delivery_event_detail_dispatch_time
  ON {TABLE}(dispatch_time);
CREATE INDEX IF NOT EXISTS idx_delivery_event_detail_sign_time
  ON {TABLE}(sign_time);
CREATE INDEX IF NOT EXISTS idx_delivery_event_detail_arrival_time
  ON {TABLE}(destination_arrival_scan_time);
CREATE INDEX IF NOT EXISTS idx_delivery_event_detail_station
  ON {TABLE}(dispatch_station_code);
CREATE INDEX IF NOT EXISTS idx_delivery_event_detail_courier
  ON {TABLE}(courier_code);
CREATE INDEX IF NOT EXISTS idx_delivery_event_detail_source
  ON {TABLE}(order_source);
"""


def load_dotenv() -> None:
    env_path = ROOT / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key, value)


def env(name: str, default: str) -> str:
    return os.environ.get(name, default)


def psql_env() -> dict[str, str]:
    values = os.environ.copy()
    values["PGPASSWORD"] = env("DATABASE_PASSWORD", "chatbi_password")
    return values


def psql_base() -> list[str]:
    return [
        "psql",
        "-h",
        env("DATABASE_HOST", "localhost"),
        "-p",
        env("DATABASE_PORT", "5432"),
        "-U",
        env("DATABASE_USER", "chatbi_user"),
        "-d",
        env("DATABASE_NAME", "chatbi_db"),
        "-v",
        "ON_ERROR_STOP=1",
    ]


def run_sql(sql: str) -> None:
    subprocess.run(psql_base(), input=sql, text=True, check=True, env=psql_env())


def scalar_sql(sql: str) -> str:
    result = subprocess.run(
        psql_base() + ["-Atc", sql],
        text=True,
        check=True,
        capture_output=True,
        env=psql_env(),
    )
    return result.stdout.strip()


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    return str(value).strip()


def parse_numeric(value):
    text = clean_text(value)
    if text == "":
        return ""
    try:
        return str(Decimal(text))
    except (InvalidOperation, ValueError):
        return ""


def parse_time(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat(sep=" ")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).isoformat(sep=" ")
    text = clean_text(value)
    if text == "":
        return ""
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt).isoformat(sep=" ")
        except ValueError:
            pass
    return ""


def parse_bool(value):
    text = clean_text(value).lower()
    if text == "":
        return ""
    if text in {"是", "yes", "y", "true", "1", "cod"}:
        return "true"
    if text in {"否", "no", "n", "false", "0", "non-cod"}:
        return "false"
    return ""


def minutes_between(start_text: str, end_text: str):
    if not start_text or not end_text:
        return ""
    try:
        start = datetime.fromisoformat(start_text)
        end = datetime.fromisoformat(end_text)
    except ValueError:
        return ""
    return str(round((end - start).total_seconds() / 60, 2))


def process_status(arrival: str, dispatch: str, sign: str) -> str:
    try:
        arrival_dt = datetime.fromisoformat(arrival) if arrival else None
        dispatch_dt = datetime.fromisoformat(dispatch) if dispatch else None
        sign_dt = datetime.fromisoformat(sign) if sign else None
    except ValueError:
        return "invalid_time"

    issues = []
    if sign_dt and not dispatch_dt:
        issues.append("signed_without_dispatch")
    if dispatch_dt and arrival_dt and dispatch_dt < arrival_dt:
        issues.append("dispatch_before_arrival")
    if sign_dt and dispatch_dt and sign_dt < dispatch_dt:
        issues.append("sign_before_dispatch")
    if sign_dt and arrival_dt and sign_dt < arrival_dt:
        issues.append("sign_before_arrival")
    if issues:
        return ",".join(issues)
    if arrival_dt and dispatch_dt and sign_dt:
        return "complete"
    if arrival_dt and dispatch_dt and not sign_dt:
        return "dispatched_not_signed"
    if arrival_dt and not dispatch_dt and not sign_dt:
        return "arrived_only"
    if dispatch_dt and not sign_dt:
        return "dispatch_only"
    if sign_dt:
        return "signed_only"
    return "no_event_time"


def raw_row_json(headers, values) -> str:
    payload = {}
    for header, value in zip(headers, values):
        if isinstance(value, (datetime, date)):
            payload[header] = value.isoformat(sep=" ")
        elif isinstance(value, Decimal):
            payload[header] = str(value)
        else:
            payload[header] = value
    return json.dumps(payload, ensure_ascii=False)


def is_empty_row(row) -> bool:
    return not any(cell is not None and str(cell).strip() != "" for cell in row)


def normalize_row(source_file: str, sheet_name: str, row_no: int, headers, row):
    values = list(row) + [None] * (len(EXPECTED_HEADERS) - len(row))
    values = values[: len(EXPECTED_HEADERS)]
    dispatch_time = parse_time(values[15])
    sign_time = parse_time(values[16])
    arrival_time = parse_time(values[19])

    return [
        source_file,
        sheet_name,
        str(row_no),
        clean_text(values[0]),
        parse_numeric(values[1]),
        parse_numeric(values[2]),
        clean_text(values[3]),
        clean_text(values[4]),
        clean_text(values[5]),
        clean_text(values[6]),
        clean_text(values[7]),
        clean_text(values[8]),
        clean_text(values[9]),
        clean_text(values[10]),
        clean_text(values[11]),
        clean_text(values[12]),
        clean_text(values[13]),
        clean_text(values[14]),
        dispatch_time,
        sign_time,
        parse_bool(values[17]),
        clean_text(values[18]),
        arrival_time,
        "true" if arrival_time else "false",
        "true" if dispatch_time else "false",
        "true" if sign_time else "false",
        minutes_between(arrival_time, dispatch_time),
        minutes_between(dispatch_time, sign_time),
        minutes_between(arrival_time, sign_time),
        process_status(arrival_time, dispatch_time, sign_time),
        raw_row_json(headers, values),
    ]


def import_workbook(path: Path, empty_report: Path) -> None:
    copy_sql = (
        f"\\copy {TABLE} ({', '.join(COLUMNS)}) "
        "FROM STDIN WITH (FORMAT csv, NULL '', QUOTE '\"')"
    )
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    empty_report.parent.mkdir(parents=True, exist_ok=True)

    total_imported = 0
    total_empty = 0
    with empty_report.open("w", newline="", encoding="utf-8") as empty_file:
        empty_writer = csv.writer(empty_file)
        empty_writer.writerow(["source_file", "source_sheet", "source_row_no"])

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = [clean_text(value) for value in next(rows)]
            except StopIteration:
                continue
            if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
                raise RuntimeError(
                    f"Unexpected headers in sheet {sheet_name}: {headers[:len(EXPECTED_HEADERS)]}"
                )

            print(f"Importing sheet {sheet_name}...", flush=True)
            proc = subprocess.Popen(
                psql_base() + ["-c", copy_sql],
                stdin=subprocess.PIPE,
                text=True,
                env=psql_env(),
            )
            assert proc.stdin is not None
            writer = csv.writer(proc.stdin, lineterminator="\n")
            imported = 0
            empty = 0
            for row_no, row in enumerate(rows, start=2):
                if is_empty_row(row):
                    empty_writer.writerow([path.name, sheet_name, row_no])
                    empty += 1
                    continue
                writer.writerow(normalize_row(path.name, sheet_name, row_no, headers, row))
                imported += 1
                if imported % 100000 == 0:
                    print(
                        f"  {sheet_name}: imported {imported} rows, empty {empty} rows...",
                        flush=True,
                    )
            proc.stdin.close()
            code = proc.wait()
            if code != 0:
                raise RuntimeError(f"COPY failed for sheet {sheet_name}")
            print(f"  {sheet_name}: imported {imported}, empty {empty}", flush=True)
            total_imported += imported
            total_empty += empty

    workbook.close()
    print(f"Workbook imported rows: {total_imported}")
    print(f"Workbook empty rows excluded: {total_empty}")
    print(f"Empty-row report: {empty_report}")


def main() -> None:
    load_dotenv()
    args = [arg for arg in sys.argv[1:] if arg != "--truncate"]
    input_path = Path(args[0]) if args else DEFAULT_INPUT
    empty_report = Path(args[1]) if len(args) > 1 else DEFAULT_EMPTY_REPORT
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    print(f"Creating table {TABLE} if needed...")
    run_sql(DDL)
    if "--truncate" in sys.argv:
        print(f"Truncating {TABLE}...")
        run_sql(f"TRUNCATE TABLE {TABLE};")

    before = scalar_sql(f"SELECT COUNT(*) FROM {TABLE};")
    print(f"Rows before import: {before}")
    import_workbook(input_path, empty_report)
    after = scalar_sql(f"SELECT COUNT(*) FROM {TABLE};")
    print(f"Rows after import: {after}")


if __name__ == "__main__":
    main()
