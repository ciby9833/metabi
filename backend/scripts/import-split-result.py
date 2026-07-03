#!/usr/bin/env python3
import csv
import os
import re
import subprocess
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_DIR = ROOT / "split_result"
TABLE = "dwd.dispatcher_efficiency_detail"

COLUMNS = [
    "source_file",
    "source_date",
    "gw_actual_ship_time",
    "waybill_no",
    "agent_area_code",
    "agent_area_name",
    "station_code",
    "station_name",
    "dispatcher_id",
    "dispatcher_name",
    "sign_type",
    "billing_weight",
    "piece_count",
    "dispatch_time",
    "dispatch_batch_code",
    "batch_node_sort",
    "is_signed_timely",
    "planned_sign_time",
    "actual_sign_time",
    "delivery_attempts",
    "is_split_delivery",
    "prev_node_address",
    "prev_node_drive_distance",
    "prev_node_duration",
    "dispatch_lng_lat",
    "unload_scan_time",
]


DDL = f"""
CREATE SCHEMA IF NOT EXISTS dwd;

CREATE TABLE IF NOT EXISTS {TABLE} (
  source_file TEXT NOT NULL,
  source_date DATE,
  gw_actual_ship_time TIMESTAMPTZ,
  waybill_no TEXT,
  agent_area_code TEXT,
  agent_area_name TEXT,
  station_code TEXT,
  station_name TEXT,
  dispatcher_id TEXT,
  dispatcher_name TEXT,
  sign_type TEXT,
  billing_weight NUMERIC,
  piece_count INT,
  dispatch_time TIMESTAMPTZ,
  dispatch_batch_code TEXT,
  batch_node_sort INT,
  is_signed_timely TEXT,
  planned_sign_time TIMESTAMPTZ,
  actual_sign_time TIMESTAMPTZ,
  delivery_attempts INT,
  is_split_delivery TEXT,
  prev_node_address TEXT,
  prev_node_drive_distance NUMERIC,
  prev_node_duration NUMERIC,
  dispatch_lng_lat TEXT,
  unload_scan_time TIMESTAMPTZ
);

CREATE INDEX IF NOT EXISTS idx_dispatcher_eff_detail_source_date
  ON {TABLE}(source_date);
CREATE INDEX IF NOT EXISTS idx_dispatcher_eff_detail_dispatch_time
  ON {TABLE}(dispatch_time);
CREATE INDEX IF NOT EXISTS idx_dispatcher_eff_detail_waybill_no
  ON {TABLE}(waybill_no);
CREATE INDEX IF NOT EXISTS idx_dispatcher_eff_detail_dispatcher
  ON {TABLE}(dispatcher_id);
CREATE INDEX IF NOT EXISTS idx_dispatcher_eff_detail_station
  ON {TABLE}(station_code);
"""


def env(name: str, default: str) -> str:
    return os.environ.get(name, default)


def load_dotenv() -> None:
    env_path = ROOT / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key, value)


def psql_base() -> list[str]:
    return [
        "psql",
        "-h",
        env("DATABASE_HOST", "localhost"),
        "-p",
        env("DATABASE_PORT", "5432"),
        "-U",
        env("DATABASE_USER", "chatbi_user"),
        "-d",
        env("DATABASE_NAME", "chatbi_db"),
        "-v",
        "ON_ERROR_STOP=1",
    ]


def run_sql(sql: str) -> None:
    env_vars = os.environ.copy()
    env_vars["PGPASSWORD"] = env("DATABASE_PASSWORD", "chatbi_password")
    subprocess.run(psql_base(), input=sql, text=True, check=True, env=env_vars)


def scalar_sql(sql: str) -> str:
    env_vars = os.environ.copy()
    env_vars["PGPASSWORD"] = env("DATABASE_PASSWORD", "chatbi_password")
    result = subprocess.run(
        psql_base() + ["-Atc", sql],
        text=True,
        check=True,
        capture_output=True,
        env=env_vars,
    )
    return result.stdout.strip()


def to_db(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float):
        return str(value)
    text = str(value).strip()
    return text


def to_int(value):
    text = to_db(value)
    if text == "":
        return ""
    try:
        return str(int(float(text)))
    except ValueError:
        return ""


def to_numeric(value):
    text = to_db(value)
    if text == "":
        return ""
    try:
        return str(Decimal(text))
    except Exception:
        return ""


def source_date_from_name(path: Path):
    match = re.match(r"(\d{2})(\d{2})", path.name)
    if not match:
        return ""
    month, day = match.groups()
    return f"2026-{month}-{day}"


def normalized_row(path: Path, row):
    values = list(row) + [None] * (24 - len(row))
    return [
        path.name,
        source_date_from_name(path),
        to_db(values[0]),
        to_db(values[1]),
        to_db(values[2]),
        to_db(values[3]),
        to_db(values[4]),
        to_db(values[5]),
        to_db(values[6]),
        to_db(values[7]),
        to_db(values[8]),
        to_numeric(values[9]),
        to_int(values[10]),
        to_db(values[11]),
        to_db(values[12]),
        to_int(values[13]),
        to_db(values[14]),
        to_db(values[15]),
        to_db(values[16]),
        to_int(values[17]),
        to_db(values[18]),
        to_db(values[19]),
        to_numeric(values[20]),
        to_numeric(values[21]),
        to_db(values[22]),
        to_db(values[23]),
    ]


def import_file(path: Path) -> None:
    copy_sql = (
        f"\\copy {TABLE} ({', '.join(COLUMNS)}) "
        "FROM STDIN WITH (FORMAT csv, NULL '', QUOTE '\"')"
    )
    env_vars = os.environ.copy()
    env_vars["PGPASSWORD"] = env("DATABASE_PASSWORD", "chatbi_password")
    proc = subprocess.Popen(
        psql_base() + ["-c", copy_sql],
        stdin=subprocess.PIPE,
        text=True,
        env=env_vars,
    )
    assert proc.stdin is not None
    writer = csv.writer(proc.stdin, lineterminator="\n")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(min_row=2, values_only=True)
    count = 0
    for row in rows:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        writer.writerow(normalized_row(path, row))
        count += 1
        if count % 50000 == 0:
            print(f"  {path.name}: {count} rows...", flush=True)

    proc.stdin.close()
    code = proc.wait()
    workbook.close()
    if code != 0:
        raise RuntimeError(f"psql COPY failed for {path}")
    print(f"  {path.name}: imported {count} rows", flush=True)


def main() -> None:
    load_dotenv()
    input_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT_DIR
    files = [
        path
        for path in sorted(input_dir.glob("*.xlsx"))
        if not path.name.startswith("~$")
    ]
    if not files:
        raise SystemExit(f"No .xlsx files found in {input_dir}")

    print(f"Creating table {TABLE} if needed...")
    run_sql(DDL)

    if "--truncate" in sys.argv:
        print(f"Truncating {TABLE}...")
        run_sql(f"TRUNCATE TABLE {TABLE};")

    before = scalar_sql(f"SELECT COUNT(*) FROM {TABLE};")
    print(f"Rows before import: {before}")
    for path in files:
        print(f"Importing {path}...")
        import_file(path)
    after = scalar_sql(f"SELECT COUNT(*) FROM {TABLE};")
    print(f"Rows after import: {after}")


if __name__ == "__main__":
    main()
